# -*- coding: utf-8 -*-
"""
ESDM Minerba Geoportal(https://geoportal.esdm.go.id/minerba/) WIUP 조회 공용 라이브러리.

CLI 스크립트(.claude/skills/esdm-wiup-lookup/scripts/search_wiup.py)와
Streamlit 앱(app.py)이 이 모듈 하나를 공유한다. 데이터 소스, 필드 매핑,
GeoJSON/엑셀 변환 로직의 단일 진실 공급원(single source of truth).
"""
from __future__ import annotations

import json
import re
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

BASE_URL = (
    "https://geoportal.esdm.go.id/monaresia/sharing/servers/"
    "3b305b4113384b41b7490479e0702093/rest/services/Pusat/WIUP_Publish/MapServer/0/query"
)

MAX_RECORD_COUNT = 100  # 서버 설정값 (MapServer capabilities에 명시됨)

# (필드명, 한글 라벨) - 엑셀/화면 표시 순서
FIELD_ORDER: list[tuple[str, str]] = [
    ("nama_usaha", "회사명(Nama Usaha)"),
    ("kode_wiup", "Single ID"),
    ("jenis_izin", "허가 종류(Jenis Izin)"),
    ("badan_usaha", "사업체 형태(Badan Usaha)"),
    ("komoditas", "광물 종류(Komoditas)"),
    ("nama_prov", "주(Provinsi)"),
    ("nama_kab", "군/시(Kabupaten)"),
    ("pulau", "섬(Pulau)"),
    ("lokasi", "상세 위치(Lokasi Tambang)"),
    ("luas_sk", "면적(Ha)"),
    ("kegiatan", "생산단계(Tahapan Kegiatan)"),
    ("cnc", "C&C 상태"),
    ("sk_iup", "허가서(SK) 번호"),
    ("pejabat", "허가권자(Pejabat)"),
    ("kode_golongan", "광물 코드"),
    ("kode_jnskom", "광물 유형 코드"),
    ("kode_wil", "지역 코드"),
]
FIELD_LABELS: dict[str, str] = dict(FIELD_ORDER)

# 검색 대상으로 고를 수 있는 필드 (필드명 -> 한글 설명)
SEARCHABLE_FIELDS: dict[str, str] = {
    "nama_usaha": "회사명",
    "lokasi": "위치 설명",
    "kode_wiup": "Single ID",
    "nama_kab": "군/시",
    "nama_prov": "주",
    "sk_iup": "SK 번호",
    "komoditas": "광물 종류",
    "jenis_izin": "허가 종류",
    "kegiatan": "생산단계",
}


# 자주 검색되는 광물의 한글명 -> 인도네시아어 표기 예시 (komoditas 필드 검색용 힌트).
# 2026-08-19에 komoditas distinct 값(163개, 대소문자 혼재) 전수 확인 후,
# 실제 존재하는 값에 대응하는 항목만 선별함. 검색은 LIKE(부분일치)+대소문자 무시라
# 아래 표기 그대로 입력하면 대소문자와 무관하게 매칭된다.
MINERAL_SEARCH_EXAMPLES: list[tuple[str, str]] = [
    ("니켈", "nikel"),
    ("보크사이트", "bauksit"),
    ("석탄", "batubara"),
    ("금", "emas"),
    ("철(철광석)", "besi"),
    ("구리", "tembaga"),
    ("주석", "timah"),
    ("납", "timbal"),
    ("망간", "mangan"),
    ("크로마이트(크롬)", "kromit"),
    ("규사(석영사)", "pasir kuarsa"),
    ("석회석", "gamping"),
    ("화강암", "granit"),
    ("안산암", "andesit"),
    ("점토", "tanah liat"),
    ("대리석", "marmer"),
    ("인회석(인산염)", "fosfat"),
    ("제올라이트", "zeolit"),
    ("흑연", "grafit"),
    ("지르콘", "zirkon"),
    ("희토류", "logam tanah jarang"),
]

# jenis_izin(허가 종류) 필드의 실제 distinct 값 전부 (2026-08-19 기준 7개, 전수 확인됨).
# 이 필드는 원래 짧은 약어라 그대로 검색어로 입력하면 됨 (예: "IUP").
PERMIT_TYPE_EXAMPLES: list[tuple[str, str]] = [
    ("WIUP (광업사업허가지역)", "WIUP"),
    ("IUP (광업사업허가)", "IUP"),
    ("IUPK (특별광업사업허가)", "IUPK"),
    ("WIUPK (특별광업허가지역)", "WIUPK"),
    ("PKP2B (석탄사업협약)", "PKP2B"),
    ("KK (계약광업권)", "KK"),
    ("IPR (민간·소규모 채굴허가)", "IPR"),
]

# kegiatan(생산단계) 필드의 실제 distinct 값 전부 (2026-08-19 기준 7개, 전수 확인됨).
PRODUCTION_STAGE_EXAMPLES: list[tuple[str, str]] = [
    ("예비지정", "PENCADANGAN"),
    ("탐사", "EKSPLORASI"),
    ("생산운영", "OPERASI PRODUKSI"),
    ("경매(입찰) 대상", "LELANG"),
    ("지원구역", "WIL. PENUNJANG"),
    ("제재(효력정지)", "SANKSI (PEMBEKUAN)"),
    ("제재(일시중단)", "SANKSI (PENGHENTIAN SEMENTARA)"),
]


class WiupApiError(RuntimeError):
    pass


def build_where(term: str, field: str = "nama_usaha", exact: bool = False) -> str:
    escaped = term.replace("'", "''")
    if field == "all":
        parts = []
        for f in ("nama_usaha", "lokasi", "kode_wiup"):
            if exact:
                parts.append(f"UPPER({f}) = UPPER('{escaped}')")
            else:
                parts.append(f"UPPER({f}) LIKE UPPER('%{escaped}%')")
        return " OR ".join(parts)
    if exact:
        return f"UPPER({field}) = UPPER('{escaped}')"
    return f"UPPER({field}) LIKE UPPER('%{escaped}%')"


def fetch_all(where: str, out_fields: str = "*", geometry: bool = True) -> list[dict]:
    """ArcGIS REST 쿼리를 resultOffset으로 페이지네이션하며 전건 수집."""
    features: list[dict] = []
    offset = 0
    while True:
        params = {
            "where": where,
            "outFields": out_fields,
            "returnGeometry": "true" if geometry else "false",
            "f": "json",
            "resultOffset": offset,
            "resultRecordCount": MAX_RECORD_COUNT,
        }
        url = BASE_URL + "?" + urllib.parse.urlencode(params)
        try:
            with urllib.request.urlopen(url, timeout=30) as resp:
                data = json.load(resp)
        except urllib.error.URLError as e:
            raise WiupApiError(f"서버 요청 실패: {e}") from e

        if "error" in data:
            raise WiupApiError(f"ArcGIS 서버 에러: {data['error']}")

        batch = data.get("features", [])
        features.extend(batch)

        if not data.get("exceededTransferLimit") or not batch:
            break
        offset += len(batch)

    return features


def search(term: str, field: str = "nama_usaha", exact: bool = False, geometry: bool = True) -> list[dict]:
    where = build_where(term, field, exact)
    return fetch_all(where, geometry=geometry)


def epoch_ms_to_date(ms: Any) -> str | None:
    if ms is None:
        return None
    try:
        return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
    except (TypeError, ValueError, OSError):
        return None


def safe_filename(name: str, fallback: str = "wiup") -> str:
    name = (name or fallback).strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", "_", name)
    return name[:120] or fallback


def features_to_geojson_dict(features: list[dict]) -> dict:
    geo_features = []
    for f in features:
        attrs = f.get("attributes", {})
        geom = f.get("geometry")
        geometry = None
        if geom is not None:
            rings = geom.get("rings")
            if rings is not None:
                geometry = {"type": "Polygon", "coordinates": rings}
        geo_features.append({"type": "Feature", "properties": attrs, "geometry": geometry})

    return {
        "type": "FeatureCollection",
        "crs": {"type": "name", "properties": {"name": "EPSG:4326"}},
        "features": geo_features,
    }


def geojson_bytes(features: list[dict]) -> bytes:
    fc = features_to_geojson_dict(features)
    return json.dumps(fc, ensure_ascii=False, indent=2).encode("utf-8")


def geojson_zip_bytes(features: list[dict]) -> bytes:
    """검색 결과 각 광구를 개별 GeoJSON 파일(`{회사명}_IUP_boundary.geojson`)로 만들어 zip으로 묶는다."""
    buf = BytesIO()
    used_names: dict[str, int] = {}
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in features:
            attrs = f.get("attributes", {})
            base_name = safe_filename(attrs.get("nama_usaha") or attrs.get("kode_wiup"))
            count = used_names.get(base_name, 0)
            used_names[base_name] = count + 1
            suffix = f"_{count + 1}" if count else ""
            file_name = f"{base_name}{suffix}_IUP_boundary.geojson"
            zf.writestr(file_name, geojson_bytes([f]))
    return buf.getvalue()


def xlsx_bytes(attributes: dict) -> bytes:
    """광구 정보(속성) 한 건을 단일 시트 엑셀로 변환 (경계좌표는 포함하지 않음 -> GeoJSON 별도 사용)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "WIUP 정보"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws["A1"] = "항목"
    ws["B1"] = "값"
    ws["A1"].font = bold
    ws["B1"].font = bold
    ws["A1"].fill = header_fill
    ws["B1"].fill = header_fill

    row = 2
    for key, label in FIELD_ORDER:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=attributes.get(key))
        row += 1

    ws.cell(row=row, column=1, value="허가 시작일(Tanggal Berlaku)")
    ws.cell(row=row, column=2, value=epoch_ms_to_date(attributes.get("tgl_berlaku")))
    row += 1
    ws.cell(row=row, column=1, value="허가 만료일(Tanggal Akhir)")
    ws.cell(row=row, column=2, value=epoch_ms_to_date(attributes.get("tgl_akhir")))
    row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_bytes_multi(features: list[dict]) -> bytes:
    """검색 결과 전체(여러 건)를 한 시트에 표 형태(1행=1광구)로 변환."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "WIUP 검색결과"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    columns = list(FIELD_ORDER) + [
        ("tgl_berlaku", "허가 시작일"),
        ("tgl_akhir", "허가 만료일"),
    ]

    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = bold
        cell.fill = header_fill

    for row_idx, f in enumerate(features, start=2):
        attrs = f.get("attributes", {})
        for col_idx, (key, _) in enumerate(columns, start=1):
            value = attrs.get(key)
            if key in ("tgl_berlaku", "tgl_akhir"):
                value = epoch_ms_to_date(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, (key, label) in enumerate(columns, start=1):
        width = max(14, min(40, len(label) + 4))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
